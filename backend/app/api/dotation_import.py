from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from typing import List, Dict, Any
import openpyxl
from io import BytesIO
from app.db.database import get_db, get_db_cursor
from app.api.auth import get_current_user

router = APIRouter(prefix="/dotation/import-excel", tags=["Dotation Import"])


def normalize_carburant(value: str) -> str:
    """Normalize carburant value to 'gazoil' or 'essence'"""
    if not value:
        return None
    v = value.lower().strip()
    if v in ['gazoil', 'gasoil', 'diesel', 'gazole']:
        return 'gazoil'
    if v in ['essence', 'super']:
        return 'essence'
    return None


@router.post("/analyze")
async def analyze_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Analyze Excel file and return preview with validation status.
    Expected columns:
    - N° POLICE (required)
    - N° CIVIL (required if vehicle doesn't exist)
    - MARQUE (required if vehicle doesn't exist)
    - CARBURANT (required if vehicle doesn't exist)
    - KM (optional, default 0)
    - SERVICE (required)
    - NOM ET PRENOM DU BENEFICIAIRE (required)
    - QTE (required)
    - QUALITE (required - fonction)
    """
    if current_user['role'] != 'ADMIN':
        raise HTTPException(status_code=403, detail="Accès administrateur requis")
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Normalize headers (remove accents, spaces, etc.)
        header_map = {}
        for i, h in enumerate(headers):
            if h:
                h_norm = str(h).strip().upper()
                header_map[h_norm] = i
        
        # Expected columns
        required_cols = {
            'POLICE': ['N° POLICE', 'POLICE', 'N POLICE', 'Nº POLICE'],
            'CIVIL': ['N° CIVIL', 'CIVIL', 'N CIVIL', 'Nº CIVIL', 'NCIVIL'],
            'MARQUE': ['MARQUE'],
            'CARBURANT': ['CARBURANT'],
            'KM': ['KM', 'KILOMETRAGE'],
            'SERVICE': ['SERVICE'],
            'NOM': ['NOM ET PRENOM DU BENEFICIAIRE', 'NOM', 'BENEFICIAIRE', 'NOM ET PRENOM'],
            'QTE': ['QTE', 'QUANTITE', 'QUOTA'],
            'FONCTION': ['QUALITE', 'FONCTION']
        }
        
        # Find column indices
        col_indices = {}
        for key, possible_names in required_cols.items():
            for name in possible_names:
                if name in header_map:
                    col_indices[key] = header_map[name]
                    break
        
        # Check required columns
        missing = []
        for key in ['POLICE', 'SERVICE', 'NOM', 'QTE', 'FONCTION']:
            if key not in col_indices:
                missing.append(key)
        
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Colonnes manquantes dans l'Excel : {', '.join(missing)}"
            )
        
        # Parse rows
        rows_data = []
        with get_db() as conn:
            cur = get_db_cursor(conn)
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                # Extract data
                police = str(row[col_indices['POLICE']]).strip() if row[col_indices['POLICE']] else None
                if not police:
                    continue
                
                civil = str(row[col_indices['CIVIL']]).strip() if 'CIVIL' in col_indices and row[col_indices['CIVIL']] else None
                marque = str(row[col_indices['MARQUE']]).strip() if 'MARQUE' in col_indices and row[col_indices['MARQUE']] else None
                carburant_raw = str(row[col_indices['CARBURANT']]).strip() if 'CARBURANT' in col_indices and row[col_indices['CARBURANT']] else None
                carburant = normalize_carburant(carburant_raw) if carburant_raw else None
                km = int(row[col_indices['KM']]) if 'KM' in col_indices and row[col_indices['KM']] else 0
                
                service_name = str(row[col_indices['SERVICE']]).strip() if row[col_indices['SERVICE']] else None
                nom = str(row[col_indices['NOM']]).strip() if row[col_indices['NOM']] else None
                qte = float(row[col_indices['QTE']]) if row[col_indices['QTE']] else None
                fonction = str(row[col_indices['FONCTION']]).strip() if row[col_indices['FONCTION']] else None
                
                if not all([police, service_name, nom, qte, fonction]):
                    continue
                
                # Validate
                errors = []
                warnings = []
                
                # Check vehicle exists
                cur.execute("SELECT id, ncivil, marque, carburant FROM vehicule WHERE police=%s", (police,))
                vehicle = cur.fetchone()
                vehicle_id = None
                vehicle_status = "exists"
                
                if vehicle:
                    vehicle_id = vehicle['id']
                else:
                    vehicle_status = "create"
                    # Need civil, marque, carburant
                    if not civil:
                        errors.append("N° CIVIL requis pour créer véhicule")
                    if not marque:
                        errors.append("MARQUE requise pour créer véhicule")
                    if not carburant:
                        errors.append("CARBURANT requis pour créer véhicule")
                
                # Check service exists - FUZZY SEARCH
                # Handle cases like "CAB/ISS" should match "ISS" or "CABINET"
                
                # Try exact match first
                cur.execute(
                    "SELECT id FROM service WHERE nom ILIKE %s OR direction ILIKE %s",
                    (service_name, service_name)
                )
                service = cur.fetchone()
                
                # If not found and contains /, try each part
                if not service and '/' in service_name:
                    parts = [p.strip() for p in service_name.split('/')]
                    for part in parts:
                        cur.execute(
                            "SELECT id FROM service WHERE nom ILIKE %s OR direction ILIKE %s",
                            (part, part)
                        )
                        service = cur.fetchone()
                        if service:
                            print(f"  → Service trouvé avec partie: '{service_name}' -> '{part}'")
                            break
                
                # If still not found, try partial match
                if not service:
                    cur.execute(
                        "SELECT id FROM service WHERE nom ILIKE %s OR direction ILIKE %s",
                        (f'%{service_name}%', f'%{service_name}%')
                    )
                    service = cur.fetchone()
                    if service:
                        print(f"  → Service trouvé avec recherche partielle: '{service_name}'")
                
                if not service:
                    print(f"  ✗ Service introuvable: '{service_name}'")
                
                service_id = service['id'] if service else None
                service_status = "exists" if service else "not_found"
                
                if not service:
                    errors.append(f"Service '{service_name}' introuvable")
                
                # Check beneficiaire exists - FUZZY SEARCH WITH DEBUG
                # Remove MR/MME prefixes and extra spaces for better matching
                nom_clean = nom.strip()
                nom_search = nom_clean
                
                # DEBUG: Show exact bytes
                print(f"  → Recherche bénéficiaire:")
                print(f"     Excel: '{nom}' (len={len(nom)})")
                print(f"     Bytes: {nom.encode('utf-8')}")
                
                # Remove common prefixes
                for prefix in ['MR ', 'MME ', 'M. ', 'MME. ', 'MONSIEUR ', 'MADAME ']:
                    if nom_clean.upper().startswith(prefix):
                        nom_search = nom_clean[len(prefix):].strip()
                        break
                
                # Try exact match first
                cur.execute("SELECT id, nom FROM benificiaire WHERE nom ILIKE %s", (nom,))
                benef = cur.fetchone()
                
                if benef:
                    print(f"     ✓ Trouvé (exact): '{benef['nom']}'")
                    print(f"     DB Bytes: {benef['nom'].encode('utf-8')}")
                else:
                    # Show what's in DB that starts with first 3 chars
                    if len(nom_search) >= 3:
                        search_start = nom_search[:3]
                        cur.execute("""
                            SELECT id, nom
                            FROM benificiaire
                            WHERE nom ILIKE %s
                            LIMIT 5
                        """, (f'{search_start}%',))
                        similar = cur.fetchall()
                        if similar:
                            print(f"     ℹ Bénéficiaires commençant par '{search_start}':")
                            for s in similar:
                                print(f"       - DB: '{s['nom']}' (len={len(s['nom'])})")
                                print(f"         Bytes: {s['nom'].encode('utf-8')}")
                
                # If not found, try without MR/MME
                if not benef and nom_search != nom:
                    cur.execute("SELECT id, nom FROM benificiaire WHERE nom ILIKE %s", (nom_search,))
                    benef = cur.fetchone()
                    if benef:
                        print(f"     ✓ Trouvé (sans préfixe): '{benef['nom']}'")
                
                # If still not found, try partial match (contains)
                if not benef:
                    cur.execute("SELECT id, nom FROM benificiaire WHERE nom ILIKE %s", (f'%{nom_search}%',))
                    benef = cur.fetchone()
                    if benef:
                        print(f"     ✓ Trouvé (partiel): '{benef['nom']}'")
                
                if not benef:
                    print(f"     ✗ AUCUN bénéficiaire trouvé pour '{nom}'")
                
                benef_id = None
                benef_status = "exists"
                
                if benef:
                    benef_id = benef['id']
                else:
                    benef_status = "create"
                    if not service_id:
                        errors.append("Service requis pour créer bénéficiaire")
                
                row_data = {
                    'row_number': row_idx,
                    'police': police,
                    'civil': civil,
                    'marque': marque,
                    'carburant': carburant,
                    'km': km,
                    'service_name': service_name,
                    'service_id': service_id,
                    'service_status': service_status,
                    'nom': nom,
                    'qte': qte,
                    'fonction': fonction,
                    'vehicle_id': vehicle_id,
                    'vehicle_status': vehicle_status,
                    'benef_id': benef_id,
                    'benef_status': benef_status,
                    'errors': errors,
                    'warnings': warnings,
                    'valid': len(errors) == 0
                }
                
                rows_data.append(row_data)
        
        # Summary
        total = len(rows_data)
        valid = sum(1 for r in rows_data if r['valid'])
        invalid = total - valid
        
        vehicles_to_create = sum(1 for r in rows_data if r['vehicle_status'] == 'create' and r['valid'])
        benefs_to_create = sum(1 for r in rows_data if r['benef_status'] == 'create' and r['valid'])
        
        return {
            'success': True,
            'summary': {
                'total_rows': total,
                'valid_rows': valid,
                'invalid_rows': invalid,
                'vehicles_to_create': vehicles_to_create,
                'beneficiaires_to_create': benefs_to_create
            },
            'rows': rows_data
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lecture Excel : {str(e)}")


@router.post("/execute")
async def execute_import(
    request: Request,
    mois: int,
    annee: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Execute the import based on analyzed rows.
    Body: { rows: [...] } - Array of validated row objects from analyze endpoint
    """
    if current_user['role'] != 'ADMIN':
        raise HTTPException(status_code=403, detail="Accès administrateur requis")
    
    # Read JSON body
    try:
        body = await request.json()
        rows = body.get('rows')  # Get rows array directly
        
        if not rows:
            raise HTTPException(status_code=400, detail="Aucune ligne à importer")
        
        if not isinstance(rows, list):
            raise HTTPException(status_code=400, detail="Format de données invalide")
        
        print(f"[IMPORT] Début import: {len(rows)} ligne(s), mois={mois}, annee={annee}")
            
    except Exception as e:
        print(f"[IMPORT ERROR] Erreur parsing body: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erreur parsing body: {str(e)}")
    
    created_vehicles = 0
    created_benefs = 0
    created_dotations = 0
    errors = []
    warnings = []
    
    with get_db() as conn:
        cur = get_db_cursor(conn)
        
        for row in rows:
            if not row['valid']:
                continue
            
            print(f"[IMPORT] Traitement ligne {row['row_number']}: police={row['police']}, nom={row['nom']}")
            
            try:
                vehicle_id = row['vehicle_id']
                benef_id = row['benef_id']
                
                # Create vehicle if needed
                if row['vehicle_status'] == 'create':
                    try:
                        print(f"  → Création véhicule {row['police']}")
                        cur.execute("""
                            INSERT INTO vehicule (police, ncivil, marque, carburant, km, actif)
                            VALUES (%s, %s, %s, %s, %s, TRUE)
                            RETURNING id
                        """, (row['police'], row['civil'], row['marque'], row['carburant'], row['km']))
                        vehicle_id = cur.fetchone()['id']
                        created_vehicles += 1
                        print(f"  ✓ Véhicule créé: id={vehicle_id}")
                    except Exception as e:
                        print(f"  ✗ Erreur création véhicule: {str(e)}")
                        errors.append({
                            'row': row['row_number'],
                            'message': f"Erreur création véhicule {row['police']}: {str(e)}"
                        })
                        continue
                
                # Create beneficiaire if needed
                if row['benef_status'] == 'create':
                    try:
                        print(f"  → Création bénéficiaire {row['nom']}")
                        # Auto-generate matricule
                        cur.execute("SELECT COUNT(*) as cnt FROM benificiaire")
                        cnt = cur.fetchone()['cnt']
                        matricule = f"B{cnt + 1:04d}"
                        
                        cur.execute("""
                            INSERT INTO benificiaire (matricule, nom, fonction, service_id)
                            VALUES (%s, %s, %s, %s)
                            RETURNING id
                        """, (matricule, row['nom'], row['fonction'], row['service_id']))
                        benef_id = cur.fetchone()['id']
                        created_benefs += 1
                        print(f"  ✓ Bénéficiaire créé: id={benef_id}")
                    except Exception as e:
                        print(f"  ✗ Erreur création bénéficiaire: {str(e)}")
                        errors.append({
                            'row': row['row_number'],
                            'message': f"Erreur création bénéficiaire {row['nom']}: {str(e)}"
                        })
                        continue
                
                # Check if dotation already exists
                cur.execute("""
                    SELECT id FROM dotation 
                    WHERE vehicule_id=%s AND mois=%s AND annee=%s
                """, (vehicle_id, mois, annee))
                existing = cur.fetchone()
                
                if existing:
                    print(f"  ⚠ Dotation existe déjà")
                    warnings.append({
                        'row': row['row_number'],
                        'message': f"Dotation existe déjà pour véhicule {row['police']} (mois {mois}/{annee}) - ignorée"
                    })
                    continue
                
                # Create dotation
                try:
                    print(f"  → Création dotation: vehicule_id={vehicle_id}, benef_id={benef_id}, qte={row['qte']}")
                    cur.execute("""
                        INSERT INTO dotation (
                            vehicule_id, benificiaire_id, mois, annee, 
                            qte, qte_consomme, cloture
                        )
                        VALUES (%s, %s, %s, %s, %s, 0, FALSE)
                        RETURNING id
                    """, (vehicle_id, benef_id, mois, annee, row['qte']))
                    
                    created_dotations += 1
                    print(f"  ✓ Dotation créée")
                except Exception as e:
                    print(f"  ✗ Erreur création dotation: {str(e)}")
                    errors.append({
                        'row': row['row_number'],
                        'message': f"Erreur création dotation: {str(e)}"
                    })
                    continue
                
            except Exception as e:
                print(f"  ✗ Erreur inattendue: {str(e)}")
                errors.append({
                    'row': row['row_number'],
                    'message': f"Erreur inattendue: {str(e)}"
                })
        
        if errors:
            conn.rollback()
            print(f"[IMPORT] ✗ Échec: {len(errors)} erreur(s)")
            for err in errors:
                print(f"  - Ligne {err['row']}: {err['message']}")
            return {
                'success': False,
                'errors': errors,
                'message': f"{len(errors)} erreur(s) - import annulé"
            }
        
        conn.commit()
        print(f"[IMPORT] ✓ Succès: {created_dotations} dotation(s), {created_vehicles} véhicule(s), {created_benefs} bénéficiaire(s)")
        
        return {
            'success': True,
            'created': {
                'vehicles': created_vehicles,
                'beneficiaires': created_benefs,
                'dotations': created_dotations
            },
            'warnings': warnings,
            'message': f"Import réussi : {created_dotations} dotation(s) créée(s)"
        }